import openpyxl


class Speed_Torque:
    def __init__(self, file=str):
        self.file = file

    def get_rpm_torque_list(self):
        dataframe = (openpyxl.load_workbook(self.file)).active
        all_list = []
        for row in range(0, dataframe.max_row):
            for col in dataframe.iter_cols(1, dataframe.max_column):
                val = str(col[row].value)
                if val != "None":
                    all_list.append(val)
        if len(all_list) % 2 == 1:
            del all_list[-1]
        return all_list

    # def seperate(self):
    #     all_list = self.get_rpm_torque_list()
    #     new_list = []
    #     for i in all_list:
    #         i = i.split()

    def sep_vals(self):
        all_list = self.get_rpm_torque_list()
        new_all = []
        for i in all_list:
            new_all.append(i.split())
        return new_all

    def out(self):
        all_list = self.sep_vals()
        new_all = []
        for item in all_list:
            new_all.append(item[0])
        speed_list = []
        torque_list = []
        a = 0
        for item in new_all:
            if a % 2 == 0:
                speed_list.append(item)
            else:
                torque_list.append(item)
            a += 1
        last = [speed_list, torque_list]
        return last

    # def sep_vals(self):
    #     all_list = self.get_rpm_torque_list()
    #     new_all = []
    #     for item in all_list:
    #         numbers = []
    #         for letter in item:
    #             if letter.isnumeric():
    #                 numbers.append(letter)
    #         new_all.append(numbers)
    #     return new_all
